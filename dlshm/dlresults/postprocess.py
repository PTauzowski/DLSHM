import glob
import os.path

import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter


def prepare_excel_multiaugmented_results(TASK_PATH,MODEL_NAME,augmentations,nrows):
    excel_pathname=os.path.join(TASK_PATH, 'ICSHM_results.xlsx')
    if os.path.exists(excel_pathname):
        # Load existing workbook
        writer = pd.ExcelWriter(excel_pathname, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    else:
        # Create new Excel file
        writer = pd.ExcelWriter(excel_pathname, engine='openpyxl')

    start_row=1
    start_col=0
    for name, postfix, fn in augmentations:
        dircontent=os.path.join(TASK_PATH,MODEL_NAME+postfix)
        excel_results = glob.glob(dircontent + '/ExcelResults/*')
        result=excel_results[0]
        df = pd.read_excel(result,skiprows=8, nrows=nrows+1, engine='openpyxl')
        df.iat[0, 0] = name
        #writer cell(row=start_row, column=0).value = name
        df.to_excel(writer, index=False, header=True if start_row == 0 else False,
                    startrow=start_row+1, startcol=start_col, sheet_name=MODEL_NAME )

        if start_col == 0:
            start_col = 10
        else:
            start_col = 0
            start_row += nrows + 2

    writer.close()

    tbr=6
    tbc=20

    source_row = 6
    source_col = 'H'
    tbi=0

    wb = load_workbook(excel_pathname)
    ws = wb[MODEL_NAME]
    #ws.cell(row=1, column=1).value = MODEL_NAME

    for row in range(tbr, tbr + nrows - 2):
        ws.cell(row=row, column=tbc).value = f'=A{row}'
    for idx, aug in enumerate(augmentations):
        ws.cell(row=tbr-1, column=20+idx+1).value=aug[0]
        for row in range(tbr, tbr+nrows-2):
            ws.cell(row=row, column=tbc+1+tbi).value = '='+source_col+f'{source_row+row-tbr}*100'
        tbi+=1
        if source_col == 'H':
            source_col = 'R'
        else:
            source_col = 'H'
            source_row += nrows + 2

    # Save changes
    wb.save(excel_pathname)

